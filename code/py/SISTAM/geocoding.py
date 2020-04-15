import os

import time

from mapbox import Geocoder

from openpyxl import workbook  # pip install openpyxl
from openpyxl import load_workbook

if __name__ == '__main__':
    start_time = time.time()

    path = os.path.join(os.path.dirname(os.path.realpath(__file__)), 'excels/direcciones-cruzados-v2.xlsx')

    wb = load_workbook(path)
    sheets = wb.sheetnames

    geocoder = Geocoder(
        access_token='pk.eyJ1IjoiZWRnYXJkb3MiLCJhIjoiY2o0OG54eDFjMGFjMzJxcGNlZzBnZnJrMiJ9.B__rQ4ztUAy6hXN7fjJHPw')

# print(sheets)

sheet = wb.active

longitude = None  # AA
latitude = None  # AB

for i in range(40000, 43059):

    Y = sheet['Y' + str(i)].value
    Z = sheet['Z' + str(i)].value

    street = Y
    number = Z

    if street is not None and number is not None:

        # Geocoding
        bbox = [-71.75320351681594, -33.15027032046216, -71.53901815869094, -33.01173911925241]

        response = geocoder.forward(number + ', ' + street, bbox=bbox, types=('address',))

        collection = response.json()

        print('i: ' + str(i))
        print(collection)

        if len(collection['features']) > 0:

            print(collection['features'])
            print(collection['features'][0])
            print(collection['features'][0]['geometry'])
            print(collection['features'][0]['geometry']['coordinates'])

            longitude = collection['features'][0]['geometry']['coordinates'][0]
            latitude = collection['features'][0]['geometry']['coordinates'][1]

        else:
            longitude = 0.0
            latitude = 0.0

        # sheet['Z2'] = response
        #
        # wb.save(path)

        sheet['AA' + str(i)] = longitude
        sheet['AB' + str(i)] = latitude

        print('Street: ' + str(street))
        print('Number: ' + str(number))
        print('Longitude: ' + str(longitude))
        print('Latitude: ' + str(latitude))
        print('**********')

wb.save(path)

print("--- %s seconds ---" % (time.time() - start_time))
