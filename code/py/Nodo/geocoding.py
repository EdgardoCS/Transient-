import pandas as pd

import os

import time

import random
import matplotlib.pyplot as plt

from mapbox import Geocoder
import datetime

from openpyxl import workbook
from openpyxl import load_workbook


if __name__ == '__main__':

    start_time = time.time()

    # Data from mixed excel
    # path = os.path.join(os.path.dirname(os.path.realpath(__file__)), 'mix-final-fonasa-avis-corregido-v3.xlsx')
    path = os.path.join(os.path.dirname(os.path.realpath(__file__)), 'mix-final-fonasa-avis-corregido-v5.xlsx')

    # Pandas
    validated = pd.read_excel(path, sheet_name='Sheet1')

    # Openpyxl
    wb = load_workbook(path)
    sheets = wb.sheetnames

    sheet = wb.active

    # # Other variables
    # one_percent = round(len(validated)/100)
    # random_values = random.sample(range(len(validated)), one_percent)   # uniform distribution

    # API Edgardo
    # geocoder = Geocoder(
    #     access_token='pk.eyJ1IjoiZWRnYXJkb3MiLCJhIjoiY2o0OG54eDFjMGFjMzJxcGNlZzBnZnJrMiJ9.B__rQ4ztUAy6hXN7fjJHPw')

    # API Alvaro
    # geocoder = Geocoder(
    #     access_token='pk.eyJ1IjoiYWx2YXJvamFyYWRpYXoiLCJhIjoiY2ltNHY1ZTQyMDFoMnUxbTN0cmpmcm83YSJ9.Wt7eHmDdX8xfCIX0UyT3Vw')

    # API Pablo
    geocoder = Geocoder(
        access_token='pk.eyJ1IjoicGFibG9yZXllczIwMjAiLCJhIjoiY2s4eDZhZmp2MTQzZjNtbWg4ZHdpYWkxYyJ9.VbRdEZ6-U2Wnsw9z3mJ1Iw')

    longitude = None  # AG
    latitude = None   # AH

    # plt.plot(random_values)
    # plt.ylabel('some numbers')
    # plt.axis([0, one_percent, 0, len(validated)])

    centers = {'CENTRO DE SALUD PADRE DAMIAN DE MOLOKAI': 'padreDamian',
               'CENTRO DE SALUD BARON': 'baron',
               'CENTRO DE SALUD LAS CAÑAS': 'lasCanas',
               'CENTRO DE SALUD REINA ISABEL II': 'reinaIsabel',
               'CENTRO DE SALUD LOS PLACERES': 'placeres',
               'CENTRO DE SALUD PUERTAS NEGRAS': 'puertasNegras',
               'CENTRO DE SALUD PLACILLA': 'placilla',
               'CENTRO DE SALUD RODELILLO': 'rodelillo',
               'CENTRO DE SALUD MARCELO MENA': 'mena',
               'CENTRO DE SALUD CORDILLERA': 'cordillera',
               'CENTRO DE SALUD ESPERANZA': 'esperanza',
               'CENTRO DE SALUD QUEBRADA VERDE': 'quebradaVerde',
               'CENTRO DE SALUD LAGUNA VERDE': 'lagunaVerde',
               'CECOSF PORVENIR BAJO': None}

    # Simple
    # bounding_boxes = {
    #     "rodelillo": [-71.5947295680045, -33.0686030854102, -71.5570823512442, -33.0458962518926],
    #     "jeanMarie": [-71.6177754269307, -33.0623164378224, -71.5993195972283, -33.0466961625223],
    #     "esperanza": [-71.5891676843287, -33.0448011549264, -71.5740074036111, -33.0288820250819],
    #     "cordillera": [-71.6486190031366, -33.0545922410214, -71.6304749450623, -33.0371169733967],
    #     "mena": [-71.6427063435407, -33.0681987326286, -71.6155850509611, -33.0433230043208],
    #     "puertasNegras": [-71.6487621244162, -33.0628082125189, -71.6366455968927, -33.0521735777052],
    #     "quebradaVerde": [-71.6584049408755, -33.0570708392714, -71.6377662756376, -33.0274965438442],
    #     "plazaJusticia": [-71.6485804175869, -33.0500567202079, -71.6018306762221, -33.0183851188857],
    #     "placeres": [-71.5984568875405, -33.0537479956391, -71.5668426315654, -33.0332317851935],
    #     "padreDamian": [-71.5757811533966, -33.0610856343429, -71.5575313055976, -33.0500376023911],
    #     "placilla": [-71.5861144109062, -33.135331999993156, -71.54991767928416, -33.10791243796176],
    #     "lagunaVerde": [-71.73896158287319, -33.14141509703812, -71.65185131353614, -33.09026985915155],
    #     "reinaIsabel": [-71.602219526485, -33.0812425059198, -71.5781048426756, -33.0501539245454],
    #     "lasCanas": [-71.6137379220317, -33.0697820210308, -71.604828804303, -33.0578739827023],
    #     "baron": [-71.6045718339726, -33.0529419698173, -71.5894638646254, -33.0376815100565]
    # }

    # Expanded
    bounding_boxes = {
        "placeres": [-71.6045718339726, -33.0686030854102, -71.5570823512442, -33.0288820250819],
        "esperanza": [-71.5984568875405, -33.0537479956391, -71.5668426315654, -33.0288820250819],
        "padreDamian": [-71.5984568875405, -33.0686030854102, -71.5570823512442, -33.0332317851935],
        "rodelillo": [-71.6045718339726, -33.0812425059198, -71.5570823512442, -33.0376815100565],
        "jeanMarie": [-71.6485804175869, -33.0812425059198, -71.5781048426756, -33.0183851188857],
        "reinaIsabel": [-71.6177754269307, -33.0812425059198, -71.5781048426756, -33.0376815100565],
        "baron": [-71.6485804175869, -33.0812425059198, -71.5781048426756, -33.0183851188857],
        "lasCanas": [-71.6177754269307, -33.0697820210308, -71.5993195972283, -33.0466961625223],
        "mena": [-71.6487621244162, -33.0681987326286, -71.5993195972283, -33.0183851188857],
        "plazaJusticia": [-71.6584049408755, -33.0681987326286, -71.5894638646254, -33.0183851188857],
        "placilla": [-71.5861144109062, -33.135331999993156, -71.54991767928416, -33.10791243796176],
        "lagunaVerde": [-71.73896158287319, -33.14141509703812, -71.65185131353614, -33.09026985915155],
        "cordillera": [-71.6584049408755, -33.0681987326286, -71.6018306762221, -33.0183851188857],
        "puertasNegras": [-71.6584049408755, -33.0681987326286, -71.6155850509611, -33.0274965438442],
        "quebradaVerde": [-71.6584049408755, -33.0628082125189, -71.6018306762221, -33.0183851188857]
    }

    georeferenced = 0
    no_georeferenced = 0
    no_information = 0
    greater_than_dot_5 = 0

    # for i in range(0, one_percent):
    # for i in range(0, len(validated)):
    for i in range(34000, 40000):     # 33331

        print("i: {_i}".format(_i=i))

        try:

            _findIt = validated.iloc[i]

            rut = _findIt['avis_document']          # .split('-')[0]
            center = _findIt['avis_center']
            street = str(_findIt['calle_corregida'])
            number = str(_findIt['numero_corregido'])

            if street is not None and number is not None:

                if center in centers:
                    bbox = bounding_boxes[centers[center]]
                else:
                    # bbox valparaiso
                    bbox = [-71.75320351681594, -33.15027032046216, -71.53901815869094, -33.01173911925241]

                response = geocoder.forward(number + ', ' + street, bbox=bbox, types=('address',))

                collection = response.json()

                if 'features' in collection:

                    relevance = None

                    if len(collection['features']) > 0:

                        longitude = collection['features'][0]['geometry']['coordinates'][0]
                        latitude = collection['features'][0]['geometry']['coordinates'][1]
                        relevance = str(collection['features'][0]['relevance'])
                        georeferenced += 1

                        if float(relevance) >= 0.5:
                            greater_than_dot_5 += 1

                    else:
                        longitude = 0.0
                        latitude = 0.0
                        no_georeferenced += 1

                    sheet['AG' + str(i + 2)] = longitude
                    sheet['AH' + str(i + 2)] = latitude
                    sheet['AI' + str(i + 2)] = ' '.join(map(str, bbox))
                    sheet['AJ' + str(i + 2)] = str(collection)
                    sheet['AK' + str(i + 2)] = str(collection['features'])
                    sheet['AL' + str(i + 2)] = relevance

            else:
                no_information += 1

        except Exception as e:
            print("Error: {_error}".format(_error=e))
            continue

        print('----------------')

    print('Saving data ...')

    wb.save(path)

    print('Done\n')

    print('Georeferenced: {_georeferenced}'.format(_georeferenced=georeferenced))
    print('No georeferenced: {_no_georeferenced}'.format(_no_georeferenced=no_georeferenced))
    print('No information: {_no_information}'.format(_no_information=no_information))
    print('Greater than 0.5: {_greater_than}'.format(_greater_than=greater_than_dot_5))

    print("--- %s seconds ---" % (time.time() - start_time))

